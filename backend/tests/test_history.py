import datetime
import io
import uuid
from decimal import Decimal
from zoneinfo import ZoneInfo

import pytest
import openpyxl
from fastapi.testclient import TestClient
from pypdf import PdfReader

from app.database import SessionLocal
from app.main import app
from app.models.bill import Bill
from app.models.dining_session import DiningSession
from app.models.menu import MenuCategory, MenuItem
from app.models.order import Order, OrderItem, OrderStatusHistory
from app.models.quick_sale import QuickSale, QuickSaleItem
from app.models.restaurant import Restaurant
from app.models.restaurant_table import RestaurantTable
from app.models.staff_user import StaffUser
from app.utils.auth import hash_password
from tests.auth_helpers import create_session_access_token as create_access_token


client = TestClient(app)


@pytest.fixture(scope="module")
def history_data():
    db = SessionLocal()
    db.query(Restaurant).filter(Restaurant.slug.in_(["history-test", "history-other"])).delete()
    db.commit()

    restaurant = Restaurant(name="History Test", slug="history-test", is_active=True, timezone="Asia/Kolkata", logo_url="/missing/history-logo.png")
    other = Restaurant(name="History Other", slug="history-other", is_active=True, timezone="Asia/Kolkata")
    db.add_all([restaurant, other])
    db.flush()

    owner = StaffUser(restaurant_id=restaurant.id, name="Owner", email="owner@history.test", password_hash=hash_password("Owner123!"), role="owner", is_active=True)
    admin = StaffUser(restaurant_id=restaurant.id, name="Admin", email="admin@history.test", password_hash=hash_password("Admin123!"), role="admin", is_active=True)
    staff = StaffUser(restaurant_id=restaurant.id, name="Server", email="server@history.test", password_hash=hash_password("Server123!"), role="staff", is_active=True)
    kitchen = StaffUser(restaurant_id=restaurant.id, name="Kitchen", email="kitchen@history.test", password_hash=hash_password("Kitchen123!"), role="kitchen", is_active=True)
    other_owner = StaffUser(restaurant_id=other.id, name="Other", email="other@history.test", password_hash=hash_password("Other123!"), role="owner", is_active=True)
    db.add_all([owner, admin, staff, kitchen, other_owner])
    db.flush()

    table = RestaurantTable(restaurant_id=restaurant.id, table_number="1", table_code="H1", is_active=True)
    pending_table = RestaurantTable(restaurant_id=restaurant.id, table_number="2", table_code="H2", is_active=True)
    other_table = RestaurantTable(restaurant_id=other.id, table_number="9", table_code="H9", is_active=True)
    db.add_all([table, pending_table, other_table])
    db.flush()

    category = MenuCategory(restaurant_id=restaurant.id, name_en="Main", display_order=1, is_active=True)
    item = MenuItem(restaurant_id=restaurant.id, category=category, name_en="Dosa", price=Decimal("100.00"), is_available=True)
    db.add_all([category, item])
    db.flush()

    # Keep fixture records inside the restaurant's local calendar day even when
    # the suite runs around midnight or while UTC is still on the prior date.
    restaurant_now = datetime.datetime.now(ZoneInfo(restaurant.timezone))
    today = restaurant_now.replace(hour=12, minute=0, second=0, microsecond=0).astimezone(datetime.timezone.utc)
    yesterday = today - datetime.timedelta(days=1)

    session = DiningSession(restaurant_id=restaurant.id, table_id=table.id, public_token=uuid.uuid4().hex, status="paid", opened_at=today - datetime.timedelta(hours=1), closed_at=today)
    db.add(session)
    db.flush()

    orders = []
    for index, created_at in enumerate([today, today - datetime.timedelta(minutes=10), yesterday], start=1):
        order = Order(
            restaurant_id=restaurant.id,
            table_id=table.id,
            dining_session_id=session.id,
            order_number=f"H-{index}",
            public_token=uuid.uuid4().hex,
            status="served" if index != 3 else "rejected",
            subtotal=Decimal("100.00") * index,
            created_at=created_at,
        )
        db.add(order)
        db.flush()
        db.add(OrderItem(
            order_id=order.id, menu_item_id=item.id,
            category_id_snapshot=category.id, category_name_snapshot=category.name_en,
            item_name="Dosa", quantity=index, unit_price=Decimal("100.00"),
            total_price=Decimal("100.00") * index,
        ))
        db.add(OrderStatusHistory(order_id=order.id, old_status="pending", new_status="accepted", changed_at=created_at + datetime.timedelta(minutes=1), changed_by_staff_id=staff.id))
        db.add(OrderStatusHistory(order_id=order.id, old_status="ready", new_status=order.status, changed_at=created_at + datetime.timedelta(minutes=5), changed_by_staff_id=staff.id))
        orders.append(order)

    bill = Bill(
        restaurant_id=restaurant.id,
        dining_session_id=session.id,
        bill_number="B-H-1",
        status="paid",
        subtotal=Decimal("300.00"),
        tax_amount=Decimal("0.00"),
        discount_amount=Decimal("0.00"),
        total_amount=Decimal("300.00"),
        # Revenue belongs to the payment day, not the earlier bill issue day.
        generated_at=yesterday,
        paid_at=today,
        payment_method="counter_cash",
    )
    db.add(bill)

    pending_session = DiningSession(
        restaurant_id=restaurant.id,
        table_id=pending_table.id,
        public_token=uuid.uuid4().hex,
        status="payment_requested",
        opened_at=today - datetime.timedelta(minutes=30),
    )
    db.add(pending_session)
    db.flush()
    db.add(Bill(
        restaurant_id=restaurant.id,
        dining_session_id=pending_session.id,
        bill_number="B-H-PENDING",
        status="issued",
        subtotal=Decimal("125.00"),
        tax_amount=Decimal("0.00"),
        discount_amount=Decimal("0.00"),
        total_amount=Decimal("125.00"),
        generated_at=today,
    ))

    quick_sale = QuickSale(
        restaurant_id=restaurant.id,
        order_number="QS-H-1",
        public_token=uuid.uuid4().hex,
        idempotency_key=uuid.uuid4().hex,
        idempotency_request_hash="a" * 64,
        sale_type="takeaway",
        source="takeaway",
        status="completed",
        subtotal=Decimal("50.00"),
        total_amount=Decimal("50.00"),
        payment_method="cash",
        entered_by_staff_id=owner.id,
        entered_by_name=owner.name,
        entered_by_role=owner.role,
        paid_by_staff_id=owner.id,
        paid_by_name=owner.name,
        paid_by_role=owner.role,
        created_at=today,
        completed_at=today,
    )
    db.add(quick_sale)
    db.flush()
    db.add(QuickSaleItem(
        quick_sale_id=quick_sale.id,
        item_name="Tea",
        quantity=1,
        base_price=Decimal("50.00"),
        unit_price=Decimal("50.00"),
        total_price=Decimal("50.00"),
    ))
    prior_local_day = (restaurant_now - datetime.timedelta(days=1)).replace(
        hour=23, minute=59, second=0, microsecond=0
    ).astimezone(datetime.timezone.utc)
    boundary_sale = QuickSale(
        restaurant_id=restaurant.id,
        order_number="QS-H-BOUNDARY",
        public_token=uuid.uuid4().hex,
        idempotency_key=uuid.uuid4().hex,
        idempotency_request_hash="b" * 64,
        sale_type="takeaway",
        source="takeaway",
        status="completed",
        subtotal=Decimal("25.00"),
        total_amount=Decimal("25.00"),
        payment_method="upi",
        entered_by_staff_id=owner.id,
        entered_by_name=owner.name,
        entered_by_role=owner.role,
        paid_by_staff_id=owner.id,
        paid_by_name=owner.name,
        paid_by_role=owner.role,
        created_at=prior_local_day,
        completed_at=prior_local_day,
    )
    db.add(boundary_sale)
    db.flush()
    db.add(QuickSaleItem(
        quick_sale_id=boundary_sale.id,
        item_name="Boundary Tea",
        quantity=1,
        base_price=Decimal("25.00"),
        unit_price=Decimal("25.00"),
        total_price=Decimal("25.00"),
    ))

    other_session = DiningSession(restaurant_id=other.id, table_id=other_table.id, public_token=uuid.uuid4().hex, status="paid", opened_at=today, closed_at=today)
    db.add(other_session)
    db.flush()
    db.add(Order(restaurant_id=other.id, table_id=other_table.id, dining_session_id=other_session.id, order_number="OTHER-1", public_token=uuid.uuid4().hex, status="served", subtotal=Decimal("999.00"), created_at=today))
    db.add(Bill(
        restaurant_id=other.id,
        dining_session_id=other_session.id,
        bill_number="B-OTHER-1",
        status="paid",
        subtotal=Decimal("999.00"),
        tax_amount=Decimal("0.00"),
        discount_amount=Decimal("0.00"),
        total_amount=Decimal("999.00"),
        generated_at=today,
        paid_at=today,
        payment_method="counter_cash",
    ))
    db.commit()

    data = {
        "owner_token": create_access_token({"sub": str(owner.id), "restaurant_id": restaurant.id, "role": "owner"}),
        "admin_token": create_access_token({"sub": str(admin.id), "restaurant_id": restaurant.id, "role": "admin"}),
        "staff_token": create_access_token({"sub": str(staff.id), "restaurant_id": restaurant.id, "role": "staff"}),
        "kitchen_token": create_access_token({"sub": str(kitchen.id), "restaurant_id": restaurant.id, "role": "kitchen"}),
        "other_token": create_access_token({"sub": str(other_owner.id), "restaurant_id": other.id, "role": "owner"}),
        "staff_id": staff.id,
        "table_id": table.id,
        "item_id": item.id,
    }
    yield data

    db = SessionLocal()
    db.query(Restaurant).filter(Restaurant.slug.in_(["history-test", "history-other"])).delete()
    db.commit()
    db.close()


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _pdf_text(response) -> str:
    reader = PdfReader(io.BytesIO(response.content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def test_operational_history_owner_admin_staff(history_data):
    assert client.get("/admin/history/orders", headers=_auth(history_data["owner_token"])).status_code == 200
    assert client.get("/admin/history/orders", headers=_auth(history_data["admin_token"])).status_code == 200
    assert client.get("/admin/history/orders", headers=_auth(history_data["staff_token"])).status_code == 200
    assert client.get("/admin/history/orders", headers=_auth(history_data["kitchen_token"])).status_code == 403
    assert client.get("/admin/history/bills", headers=_auth(history_data["staff_token"])).status_code == 403
    assert client.get("/admin/history/bills/export", headers=_auth(history_data["staff_token"])).status_code == 403
    assert client.get("/admin/history/performance", headers=_auth(history_data["staff_token"])).status_code == 403


def test_today_yesterday_and_pagination(history_data):
    today = client.get("/admin/history/orders?preset=today&page=1&page_size=1", headers=_auth(history_data["owner_token"]))
    assert today.status_code == 200
    body = today.json()
    assert body["total"] == 3
    assert len(body["items"]) == 1

    yesterday = client.get("/admin/history/orders?preset=yesterday", headers=_auth(history_data["owner_token"]))
    assert yesterday.status_code == 200
    assert yesterday.json()["total"] == 2


def test_cross_restaurant_isolation(history_data):
    response = client.get("/admin/history/orders?preset=last_7_days", headers=_auth(history_data["owner_token"]))
    assert response.status_code == 200
    order_numbers = {item["order_number"] for item in response.json()["items"]}
    assert "OTHER-1" not in order_numbers


def test_performance_revenue_and_top_selling(history_data):
    response = client.get("/admin/history/performance?preset=today", headers=_auth(history_data["owner_token"]))
    assert response.status_code == 200
    body = response.json()
    assert body["metrics"]["total_revenue"] == "350.00"
    assert body["metrics"]["collected_revenue"] == "350.00"
    assert body["metrics"]["completed_quick_sale_revenue"] == "50.00"
    assert body["metrics"]["pending_collection"] == "125.00"
    assert body["metrics"]["total_orders"] == 3
    assert body["metrics"]["average_order_value"] == "175.00"
    assert body["top_selling_items"][0]["item_name"] == "Dosa"
    assert {row["item_name"] for row in body["top_selling_items"]} == {"Dosa", "Tea"}
    assert body["category_performance"] == [{"category_name": "Main", "quantity": 3, "revenue": "300.00"}]
    assert {row["table_number"] for row in body["table_usage"]} == {"1", "2"}
    assert next(row for row in body["table_usage"] if row["table_number"] == "1")["revenue"] == "300.00"
    assert next(row for row in body["table_usage"] if row["table_number"] == "2")["revenue"] == "0.00"
    assert body["sales_mix"] == [
        {"label": "Dine-in", "revenue": "300.00", "contribution_percentage": "85.71"},
        {"label": "Takeaway", "revenue": "50.00", "contribution_percentage": "14.29"},
        {"label": "Quick Sale", "revenue": "0.00", "contribution_percentage": "0.00"},
    ]
    assert "Takeaway contributed 14.3% of collected revenue." in body["owner_insights"]
    assert "Average order value was INR 175.00." in body["owner_insights"]


@pytest.mark.parametrize("preset", ["today", "last_7_days", "month"])
def test_performance_populated_presets_serialize_all_tabs(history_data, preset):
    response = client.get(f"/admin/history/performance?preset={preset}", headers=_auth(history_data["owner_token"]))
    assert response.status_code == 200
    body = response.json()
    assert set(body) == {
        "metrics",
        "revenue_by_day",
        "orders_by_day",
        "orders_by_hour",
        "top_selling_items",
        "lowest_selling_items",
        "category_performance",
        "table_usage",
        "staff_activity",
        "sales_mix",
        "owner_insights",
    }
    assert all(isinstance(body[field], list) for field in set(body) - {"metrics"})
    restaurant_today = datetime.datetime.now(ZoneInfo("Asia/Kolkata")).date()
    includes_boundary_sale = preset == "last_7_days" or (
        preset == "month" and restaurant_today.day > 1
    )
    assert body["metrics"]["total_revenue"] == ("375.00" if includes_boundary_sale else "350.00")


def test_performance_uses_restaurant_local_day_boundary(history_data):
    restaurant_today = datetime.datetime.now(ZoneInfo("Asia/Kolkata")).date()
    previous_day = restaurant_today - datetime.timedelta(days=1)
    response = client.get(
        f"/admin/history/performance?preset=custom&start_date={previous_day}&end_date={previous_day}",
        headers=_auth(history_data["owner_token"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["metrics"]["completed_quick_sale_revenue"] == "25.00"
    assert body["metrics"]["collected_revenue"] == "25.00"
    assert body["revenue_by_day"] == [{"date": str(previous_day), "revenue": "25.00"}]


def test_performance_custom_empty_period_serializes_all_tabs(history_data):
    response = client.get(
        "/admin/history/performance?preset=custom&start_date=2000-01-01&end_date=2000-01-02",
        headers=_auth(history_data["owner_token"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["metrics"]["total_revenue"] == "0.00"
    assert body["metrics"]["pending_collection"] == "0.00"
    assert all(body[field] == [] for field in set(body) - {"metrics", "sales_mix"})
    assert body["sales_mix"] == [
        {"label": "Dine-in", "revenue": "0.00", "contribution_percentage": "0.00"},
        {"label": "Takeaway", "revenue": "0.00", "contribution_percentage": "0.00"},
        {"label": "Quick Sale", "revenue": "0.00", "contribution_percentage": "0.00"},
    ]


def test_custom_range_empty_and_csv_export(history_data):
    response = client.get(
        "/admin/history/orders?preset=custom&start_date=2000-01-01&end_date=2000-01-02",
        headers=_auth(history_data["owner_token"]),
    )
    assert response.status_code == 200
    assert response.json()["total"] == 0

    csv_response = client.get("/admin/history/orders/export?preset=today", headers=_auth(history_data["owner_token"]))
    assert csv_response.status_code == 200
    assert "text/csv" in csv_response.headers["content-type"]
    assert "Order Number" in csv_response.text
    assert "order_number" not in csv_response.text


def test_history_and_gst_custom_ranges_reject_more_than_370_days(history_data):
    headers = _auth(history_data["owner_token"])
    params = "preset=custom&start_date=2024-01-01&end_date=2025-01-06"

    history_response = client.get(f"/admin/history/orders?{params}", headers=headers)
    gst_response = client.get(f"/admin/gst/summary?{params}", headers=headers)

    assert history_response.status_code == 422
    assert history_response.json()["detail"] == "Date range cannot exceed 370 days"
    assert gst_response.status_code == 422
    assert gst_response.json()["detail"] == "Date range cannot exceed 370 days"


def test_performance_pdf_owner_admin_access_and_rejections(history_data):
    owner = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["owner_token"]))
    admin = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["admin_token"]))
    staff = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["staff_token"]))
    kitchen = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["kitchen_token"]))
    public = client.get("/admin/history/performance/export.pdf?preset=today")

    assert owner.status_code == 200
    assert admin.status_code == 200
    assert staff.status_code == 403
    assert kitchen.status_code == 403
    assert public.status_code in {401, 403}


def test_performance_pdf_daily_content_type_filename_and_totals(history_data):
    summary = client.get("/admin/history/performance?preset=today", headers=_auth(history_data["owner_token"])).json()
    response = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["owner_token"]))

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.headers["content-disposition"].startswith('attachment; filename="omlu-history-test-performance-')
    text = _pdf_text(response)
    assert "OMLU Daily Performance Report" in text
    assert "History Test" in text
    assert "Selected reporting period" in text
    assert f"INR {summary['metrics']['total_revenue']}" in text
    assert "Total Orders" in text
    assert "Average Table Session" in text
    assert str(summary["metrics"]["total_orders"]) in text
    assert "Sales Mix" in text
    assert "Takeaway" in text
    assert "14.29%" in text
    assert "Owner Insights" in text


def test_performance_pdf_monthly_and_custom_date_ranges(history_data):
    monthly = client.get("/admin/history/performance/export.pdf?preset=month", headers=_auth(history_data["owner_token"]))
    custom = client.get(
        "/admin/history/performance/export.pdf?preset=custom&start_date=2000-01-01&end_date=2000-01-02",
        headers=_auth(history_data["owner_token"]),
    )

    assert monthly.status_code == 200
    assert "omlu-history-test-performance-" in monthly.headers["content-disposition"]
    assert "OMLU Monthly Performance Report" in _pdf_text(monthly)
    assert custom.status_code == 200
    assert "omlu-history-test-performance-2000-01-01-to-2000-01-02.pdf" in custom.headers["content-disposition"]
    custom_text = _pdf_text(custom)
    assert "OMLU Performance Report" in custom_text
    assert "No data available" in custom_text


def test_performance_pdf_restaurant_isolation_and_logo_fallback(history_data):
    response = client.get("/admin/history/performance/export.pdf?preset=today", headers=_auth(history_data["other_token"]))

    assert response.status_code == 200
    text = _pdf_text(response)
    assert "History Other" in text
    assert "History Test" not in text
    assert "INR 300.00" not in text


def test_performance_csv_is_spreadsheet_friendly(history_data):
    response = client.get("/admin/history/performance/export?preset=today", headers=_auth(history_data["owner_token"]))

    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert 'filename="omlu-history-test-performance-' in response.headers["content-disposition"]
    assert "Section,Measure,Value,Unit" in response.text
    assert "Executive Summary,Total Revenue,350.00,INR" in response.text
    assert "Quick Sale Revenue" in response.text
    assert "Sales Mix,Takeaway,50.00,INR" in response.text
    assert "Sales Mix,Takeaway Contribution,14.29,Percent" in response.text
    assert "total_revenue" not in response.text
    assert "completed_quick_sale_revenue" not in response.text


def test_performance_xlsx_management_report(history_data):
    response = client.get("/admin/history/performance/export.xlsx?preset=month", headers=_auth(history_data["owner_token"]))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert 'filename="omlu-history-test-performance-' in response.headers["content-disposition"]
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Management Report", "Revenue Trend"]
    report = workbook["Management Report"]
    assert report["A1"].value == "OMLU Monthly Performance Report"
    values = [cell.value for row in report.iter_rows() for cell in row]
    for heading in ("Executive Summary", "Sales Mix", "Revenue Trend", "Order Health", "Operations", "Top Performance", "Owner Insights"):
        if heading != "Revenue Trend":
            assert heading in values
    assert "total_revenue" not in " ".join(str(value) for value in values if value is not None)
    assert workbook["Revenue Trend"]._charts


def test_category_performance_and_order_history_survive_menu_item_deletion(history_data):
    headers = _auth(history_data["owner_token"])
    before = client.get("/admin/history/performance?preset=today", headers=headers).json()
    before_category = before["category_performance"]

    deleted = client.delete(f"/admin/menu-items/{history_data['item_id']}", headers=headers)
    assert deleted.status_code == 204

    orders = client.get("/admin/history/orders?preset=today", headers=headers)
    after = client.get("/admin/history/performance?preset=today", headers=headers)
    assert orders.status_code == 200
    assert orders.json()["total"] == 3
    assert {order["order_number"] for order in orders.json()["items"]} >= {"H-1", "H-2"}
    assert after.status_code == 200
    assert after.json()["category_performance"] == before_category
    assert after.json()["metrics"] == before["metrics"]
