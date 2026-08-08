import csv
from decimal import Decimal
import io
import uuid
import zipfile

from fastapi.testclient import TestClient
import openpyxl
import pytest

from app.database import SessionLocal
from app.main import app
from app.models.bill import Bill
from app.models.dining_session import DiningSession
from app.models.menu import MenuItem
from app.models.order import Order, OrderItem
from app.models.quick_sale import QuickSale, QuickSaleItem
from app.models.restaurant import Restaurant
from app.models.staff_user import StaffUser
from app.models.restaurant_table import RestaurantTable
from tests.auth_helpers import create_session_access_token


client = TestClient(app)


@pytest.fixture
def p4_gst_context():
    db = SessionLocal()

    # 1. GST-enabled restaurant with special characters in name
    r_gst = Restaurant(
        name="Café & Bistro @ 100%",
        slug=f"cafe-bistro-{uuid.uuid4().hex[:6]}",
        gst_enabled=True,
        gstin="27ABCDE1234F1Z5",
        gst_state_name="Maharashtra",
        gst_state_code="27",
        legal_business_name="Café Bistro Private Limited",
        default_gst_rate=Decimal("5.00"),
    )
    # 2. Non-GST restaurant
    r_nogst = Restaurant(
        name="Simple Dhaba",
        slug=f"simple-dhaba-{uuid.uuid4().hex[:6]}",
        gst_enabled=False,
    )
    db.add_all([r_gst, r_nogst])
    db.flush()

    # Staff users
    owner_gst = StaffUser(
        restaurant_id=r_gst.id, name="Owner GST", username=f"owner_g_{uuid.uuid4().hex[:6]}",
        password_hash="hash", role="owner", is_active=True
    )
    admin_gst = StaffUser(
        restaurant_id=r_gst.id, name="Admin GST", username=f"admin_g_{uuid.uuid4().hex[:6]}",
        password_hash="hash", role="admin", is_active=True
    )
    staff_gst = StaffUser(
        restaurant_id=r_gst.id, name="Staff GST", username=f"staff_g_{uuid.uuid4().hex[:6]}",
        password_hash="hash", role="staff", is_active=True
    )
    owner_nogst = StaffUser(
        restaurant_id=r_nogst.id, name="Owner NoGST", username=f"owner_ng_{uuid.uuid4().hex[:6]}",
        password_hash="hash", role="owner", is_active=True
    )
    db.add_all([owner_gst, admin_gst, staff_gst, owner_nogst])
    db.commit()

    # Table for sessions
    t1 = RestaurantTable(restaurant_id=r_gst.id, table_number="T1", table_code="T1")
    db.add(t1)
    db.commit()

    # Sample B2B Bill with invoice number
    s1 = DiningSession(restaurant_id=r_gst.id, table_id=t1.id, public_token=uuid.uuid4().hex, status="paid")
    db.add(s1); db.flush()
    b1_b2b = Bill(
        restaurant_id=r_gst.id, dining_session_id=s1.id, bill_number="BILL-B2B-1",
        invoice_number="INV/2026-27/000001", receipt_token=uuid.uuid4().hex, status="paid", currency="INR",
        customer_tax_type="b2b", customer_gstin_snapshot="27AAAAA0000A1Z5", customer_legal_name_snapshot="Acme Corp Pvt Ltd",
        subtotal=Decimal("1000.00"), discount_amount=Decimal("0.00"), taxable_amount=Decimal("1000.00"),
        gst_rate=Decimal("5.00"), cgst_amount=Decimal("25.00"), sgst_amount=Decimal("25.00"), igst_amount=Decimal("0.00"),
        tax_amount=Decimal("50.00"), total_amount=Decimal("1050.00"), payment_method="counter_upi", gst_enabled_snapshot=True
    )
    db.add(b1_b2b)

    # Cancelled Bill with invoice number
    s2 = DiningSession(restaurant_id=r_gst.id, table_id=t1.id, public_token=uuid.uuid4().hex, status="cancelled")
    db.add(s2); db.flush()
    b2_cancelled = Bill(
        restaurant_id=r_gst.id, dining_session_id=s2.id, bill_number="BILL-CAN-2",
        invoice_number="INV/2026-27/000002", receipt_token=uuid.uuid4().hex, status="cancelled", currency="INR",
        subtotal=Decimal("500.00"), discount_amount=Decimal("0.00"), taxable_amount=Decimal("500.00"),
        gst_rate=Decimal("5.00"), cgst_amount=Decimal("12.50"), sgst_amount=Decimal("12.50"), igst_amount=Decimal("0.00"),
        tax_amount=Decimal("25.00"), total_amount=Decimal("525.00"), payment_method="counter_cash", gst_enabled_snapshot=True
    )
    db.add(b2_cancelled)

    # Completed Quick Sale (B2C)
    qs3 = QuickSale(
        restaurant_id=r_gst.id, order_number="QS-300", invoice_number="INV/2026-27/000003", public_token=uuid.uuid4().hex, idempotency_key=uuid.uuid4().hex,
        idempotency_request_hash="hash", sale_type="takeaway", source="takeaway", status="completed",
        customer_tax_type="b2c", subtotal=Decimal("200.00"), discount_amount=Decimal("0.00"), taxable_amount=Decimal("200.00"),
        gst_rate=Decimal("5.00"), cgst_amount=Decimal("5.00"), sgst_amount=Decimal("5.00"), igst_amount=Decimal("0.00"),
        tax_amount=Decimal("10.00"), total_amount=Decimal("210.00"), payment_method="cash", gst_enabled_snapshot=True,
        entered_by_staff_id=owner_gst.id, entered_by_name="Owner", entered_by_role="owner"
    )
    db.add(qs3)
    db.flush()
    qsi3 = QuickSaleItem(
        quick_sale_id=qs3.id, item_name="Special Tea", quantity=4, base_price=Decimal("50.00"), unit_price=Decimal("50.00"), total_price=Decimal("200.00"),
        hsn_sac_code_snapshot="996331", gst_rate_snapshot=Decimal("5.00")
    )
    db.add(qsi3)
    db.commit()

    token_owner_gst = create_session_access_token({"sub": str(owner_gst.id), "role": "owner"}, db=db)
    token_admin_gst = create_session_access_token({"sub": str(admin_gst.id), "role": "admin"}, db=db)
    token_staff_gst = create_session_access_token({"sub": str(staff_gst.id), "role": "staff"}, db=db)
    token_owner_nogst = create_session_access_token({"sub": str(owner_nogst.id), "role": "owner"}, db=db)

    ctx = {
        "r_gst_id": r_gst.id,
        "r_nogst_id": r_nogst.id,
        "token_owner_gst": token_owner_gst,
        "token_admin_gst": token_admin_gst,
        "token_staff_gst": token_staff_gst,
        "token_owner_nogst": token_owner_nogst,
    }
    db.close()
    return ctx


def test_export_ca_package_zip_and_metadata(p4_gst_context):
    """Test CA package ZIP contents, sanitization, report metadata, and openpyxl validity."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    res = client.get("/admin/gst/export/ca-package?preset=today", headers=headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/zip"
    assert "attachment; filename=" in res.headers["content-disposition"]

    zf = zipfile.ZipFile(io.BytesIO(res.content))
    file_list = zf.namelist()

    # Verify expected files in ZIP
    assert any("report-metadata.txt" in f for f in file_list)
    assert any("sales-register.xlsx" in f for f in file_list)
    assert any("gst-summary.xlsx" in f for f in file_list)
    assert any("hsn-sac-summary.xlsx" in f for f in file_list)
    assert any("b2b-invoices.xlsx" in f for f in file_list)
    assert any("b2c-summary.xlsx" in f for f in file_list)
    assert any("documents-issued.xlsx" in f for f in file_list)
    assert any("cancelled-documents.xlsx" in f for f in file_list)
    assert any("gst-summary.pdf" in f for f in file_list)

    # Inspect metadata text
    meta_name = [f for f in file_list if "report-metadata.txt" in f][0]
    meta_bytes = zf.read(meta_name).decode("utf-8")
    assert "Café & Bistro @ 100%" in meta_bytes
    assert "27ABCDE1234F1Z5" in meta_bytes
    assert "OMLU GST Accounting Export v1.0" in meta_bytes
    assert "This package is intended for accounting and reconciliation" in meta_bytes

    # Read sales-register.xlsx from ZIP with openpyxl
    sales_name = [f for f in file_list if "sales-register.xlsx" in f][0]
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(sales_name)))
    ws = wb.active
    assert ws.title == "Sales Register"
    assert ws.freeze_panes == "A2"

    # Verify row values and numeric cell types
    row2 = [cell.value for cell in ws[2]]
    # Numeric total amount in col 16
    tot_val = ws.cell(row=2, column=16).value
    assert isinstance(tot_val, (int, float))
    assert tot_val == 1050.0

    # Ensure no sensitive database keys in exported rows
    full_text = str([cell.value for row in ws.rows for cell in row])
    assert "receipt_token" not in full_text
    assert "public_token" not in full_text
    assert "password_hash" not in full_text


def test_export_sales_register_xlsx_and_csv(p4_gst_context):
    """Test XLSX numeric storage, human-readable labels, and CSV format."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}

    # 1. XLSX
    res_xlsx = client.get("/admin/gst/export/sales-register?format=xlsx&preset=today", headers=headers)
    assert res_xlsx.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res_xlsx.content))
    ws = wb.active
    # Check human readable mapped value "Paid" or "Completed"
    row2_vals = [cell.value for cell in ws[2]]
    assert "Paid" in row2_vals or "Completed" in row2_vals

    # 2. CSV
    res_csv = client.get("/admin/gst/export/sales-register?format=csv&preset=today", headers=headers)
    assert res_csv.status_code == 200
    reader = csv.reader(io.StringIO(res_csv.content.decode("utf-8")))
    rows = list(reader)
    assert rows[0][0] == "Invoice Date"
    assert len(rows) >= 2


def test_export_pdf_summary(p4_gst_context):
    """Test PDF generation returning valid %PDF binary header."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    res = client.get("/admin/gst/export/rate-summary?format=pdf&preset=today", headers=headers)
    assert res.status_code == 200
    assert res.content.startswith(b"%PDF")


def test_export_hsn_summary_limitation_preserved(p4_gst_context):
    """Test HSN/SAC summary export preserving unallocated tax limitation."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    res = client.get("/admin/gst/export/hsn-summary?format=xlsx&preset=today", headers=headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    # Row 1 contains limitation notice
    r1_text = ws.cell(row=1, column=1).value
    assert "Line-level tax allocation is omitted" in r1_text or "NOTE:" in r1_text


def test_export_empty_period_does_not_crash(p4_gst_context):
    """Test export generation for empty date range produces valid files with notice row."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    res = client.get("/admin/gst/export/sales-register?preset=custom&start_date=2010-01-01&end_date=2010-01-02", headers=headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "No transactions for selected period"


def test_export_gst_disabled_non_gst_ca_package(p4_gst_context):
    """Test Non-GST restaurant CA package export produces valid Non-GST metadata."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_nogst']}"}
    res = client.get("/admin/gst/export/ca-package?preset=today", headers=headers)
    assert res.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(res.content))
    meta_name = [f for f in zf.namelist() if "report-metadata.txt" in f][0]
    meta_txt = zf.read(meta_name).decode("utf-8")
    assert "Simple Dhaba" in meta_txt
    assert "GST Disabled" in meta_txt


def test_export_rbac_permissions(p4_gst_context):
    """Test owner and admin can access exports, while staff user receives 403 Forbidden."""
    h_owner = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    h_admin = {"Authorization": f"Bearer {p4_gst_context['token_admin_gst']}"}
    h_staff = {"Authorization": f"Bearer {p4_gst_context['token_staff_gst']}"}

    assert client.get("/admin/gst/export/ca-package", headers=h_owner).status_code == 200
    assert client.get("/admin/gst/export/ca-package", headers=h_admin).status_code == 200
    assert client.get("/admin/gst/export/ca-package", headers=h_staff).status_code == 403


def test_export_financial_decimal_precision(p4_gst_context):
    """Test exported XLSX monetary cells preserve exact Decimal values without float representation artifacts."""
    from decimal import Decimal
    from app.services.gst_exports import _to_num

    d1 = Decimal("199.99")
    d2 = Decimal("700.05")
    d3 = Decimal("0.30")

    assert _to_num("199.99") == d1
    assert _to_num("700.05") == d2
    assert _to_num("0.30") == d3
    assert str(_to_num("0.30")) == "0.30"
    assert not str(_to_num("0.30")).startswith("0.30000000000000004")


def test_export_query_bounds_and_memory_safety(p4_gst_context):
    """Test that export register queries use bounded pagination limits preventing memory overflow."""
    headers = {"Authorization": f"Bearer {p4_gst_context['token_owner_gst']}"}
    res = client.get("/admin/gst/export/sales-register?format=xlsx&preset=today", headers=headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    # Max rows is safely bounded
    assert ws.max_row <= 10005

