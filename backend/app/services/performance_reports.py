import io
import re
from calendar import monthrange
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def performance_report_title(report_type: str) -> str:
    if report_type == "Monthly report":
        return "OMLU Monthly Performance Report"
    if report_type == "Daily report":
        return "OMLU Daily Performance Report"
    return "OMLU Performance Report"


def performance_filename(restaurant_name: str, start_date: date, end_date: date, extension: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", restaurant_name.casefold()).strip("-") or "restaurant"
    is_complete_month = (
        start_date.day == 1
        and end_date.year == start_date.year
        and end_date.month == start_date.month
        and end_date.day == monthrange(start_date.year, start_date.month)[1]
    )
    period = start_date.strftime("%Y-%m") if is_complete_month else (
        start_date.isoformat() if start_date == end_date else f"{start_date.isoformat()}-to-{end_date.isoformat()}"
    )
    return f"omlu-{slug}-performance-{period}.{extension}"


def owner_insights(metrics: dict[str, Any], sales_mix: list[dict[str, Any]], top_items: list[dict[str, Any]]) -> list[str]:
    insights: list[str] = []
    for row in sales_mix:
        contribution = Decimal(str(row["contribution_percentage"]))
        if contribution > 0:
            insights.append(f"{row['label']} contributed {contribution.quantize(Decimal('0.1'))}% of collected revenue.")
    if Decimal(str(metrics.get("pending_collection") or 0)) == 0 and Decimal(str(metrics.get("collected_revenue") or 0)) > 0:
        insights.append("All recorded revenue was collected; there is no pending collection.")
    if int(metrics.get("total_orders") or 0) > 0:
        average = Decimal(str(metrics.get("average_order_value") or 0)).quantize(Decimal("0.01"))
        insights.append(f"Average order value was INR {average}.")
    if top_items:
        best = max(top_items, key=lambda row: int(row.get("quantity") or 0))
        insights.append(f"{best['item_name']} was the most ordered item, with {int(best['quantity'])} sold.")
    return insights[:5]


def build_performance_xlsx(context: dict[str, Any]) -> bytes:
    summary = context["summary"]
    metrics = summary["metrics"]
    report = context["report"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Management Report"
    dark = PatternFill("solid", fgColor="1F2937")
    orange = PatternFill("solid", fgColor="F97316")
    pale = PatternFill("solid", fgColor="FFF7ED")

    ws.merge_cells("A1:D1")
    ws["A1"] = report["title"]
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = dark
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws["A2"] = "Restaurant"
    ws["B2"] = context["restaurant"]["name"]
    ws["A3"] = "Reporting Period"
    ws["B3"] = f"{report['start_date']} to {report['end_date']}"
    ws["A4"] = "Generated"
    ws["B4"] = report["generated_at"]

    def section(title: str) -> None:
        ws.append([])
        ws.append([title])
        cell = ws.cell(ws.max_row, 1)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)

    def rows(values: list[list[Any]]) -> None:
        for value in values:
            ws.append(value)

    section("Executive Summary")
    rows([
        ["Measure", "Value", "Measure", "Value"],
        ["Total Revenue", Decimal(str(metrics["total_revenue"])), "Collected Revenue", Decimal(str(metrics["collected_revenue"]))],
        ["Pending Collection", Decimal(str(metrics["pending_collection"])), "Total Orders", int(metrics["total_orders"])],
        ["Average Order Value", Decimal(str(metrics["average_order_value"])), "Total Bills", int(metrics["total_bills"])],
    ])

    section("Sales Mix")
    rows([["Sale Type", "Revenue", "Contribution (%)"]])
    for row in context["sales_mix"]:
        rows([[row["label"], Decimal(str(row["revenue"])), Decimal(str(row["contribution_percentage"]))]])

    section("Order Health")
    rows([["Cancelled Orders", "Not tracked"], ["Rejected Orders", int(metrics["rejected_orders"])], ["Payment Failures", "Not tracked"]])
    section("Operations")
    rows([["Average Table Session (Minutes)", int(metrics["average_session_duration_minutes"])], ["Active Table Time (Minutes)", int(metrics["active_table_time_minutes"])]])

    section("Top Performance")
    rows([["Item", "Quantity Sold", "Revenue"]])
    for item in sorted(summary["top_selling_items"], key=lambda row: (Decimal(str(row["revenue"])), row["quantity"]), reverse=True)[:10]:
        rows([[item["item_name"], int(item["quantity"]), Decimal(str(item["revenue"]))]])

    section("Owner Insights")
    for insight in context["owner_insights"]:
        rows([[f"• {insight}"]])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)

    trend = wb.create_sheet("Revenue Trend")
    trend.append([report["title"]])
    trend.append([context["restaurant"]["name"]])
    trend.append(["Date", "Revenue"])
    for row in summary["revenue_by_day"]:
        trend.append([date.fromisoformat(row["date"]), Decimal(str(row["revenue"]))])
    if summary["revenue_by_day"]:
        chart = LineChart()
        chart.title = "Daily Revenue"
        chart.y_axis.title = "Revenue (INR)"
        chart.x_axis.title = "Date"
        chart.add_data(Reference(trend, min_col=2, min_row=3, max_row=trend.max_row), titles_from_data=True)
        chart.set_categories(Reference(trend, min_col=1, min_row=4, max_row=trend.max_row))
        chart.height = 8
        chart.width = 16
        trend.add_chart(chart, "D3")

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A5" if sheet is ws else "A4"
        for column in "ABCD":
            sheet.column_dimensions[column].width = 26
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row > 1 and cell.value is not None and cell.fill.fill_type is None:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet is ws:
            for row in range(1, ws.max_row + 1):
                if ws.cell(row, 1).value in {"Measure", "Sale Type", "Item"}:
                    for cell in ws[row]:
                        cell.fill = pale
                        cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
