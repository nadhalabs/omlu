import datetime
from decimal import Decimal
import io
import re
from typing import Any, Dict, List, Optional, Tuple
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.staff_user import StaffUser
from app.services.gst_reports import (
    get_b2b_register,
    get_b2c_register,
    get_cancelled_documents_register,
    get_documents_issued_audit,
    get_gst_center_summary,
    get_gst_rate_summary,
    get_gst_sales_register,
    get_hsn_summary,
    resolve_gst_period_bounds,
)


DISCLAIMER_TEXT = (
    "This package is intended for accounting and reconciliation and is not direct GST filing with GSTN."
)

STATUS_MAP = {
    "cash": "Cash",
    "upi": "UPI",
    "card": "Card",
    "other": "Other",
    "counter_cash": "Cash",
    "counter_upi": "UPI",
    "counter_card": "Card",
    "online": "Online",
    "payment_pending": "Payment Pending",
    "issued": "Issued",
    "paid": "Paid",
    "completed": "Completed",
    "cancelled": "Cancelled",
    "active": "Active",
    "b2b": "GST Invoice",
    "b2c": "Regular Sale",
    "bill": "Bill",
    "quick_sale": "Quick Sale",
    "takeaway": "Takeaway",
    "late_entry": "Late Entry",
    "needs_review": "Needs Review",
    "unallocated_header_discount": "Unallocated Header Discount",
}


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w\s-]", "", name).strip()
    return re.sub(r"[-\s_]+", "-", cleaned).lower() or "restaurant"


def _map_val(val: Any) -> Any:
    if val is None:
        return "N/A"
    if isinstance(val, str) and val in STATUS_MAP:
        return STATUS_MAP[val]
    return val


def _to_num(val: Any) -> Optional[Decimal]:
    if val is None or val == "N/A" or val == "—":
        return None
    try:
        return Decimal(str(val))
    except Exception:
        return None


def generate_report_metadata(restaurant: Any, preset: Optional[str], start_local: datetime.date, end_local: datetime.date) -> str:
    is_gst = bool(restaurant.gst_enabled)
    now_utc = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    lines = [
        "================================================================================",
        "                    OMLU GST ACCOUNTING & RECONCILIATION EXPORT                  ",
        "================================================================================",
        f"Version Marker:        OMLU GST Accounting Export v1.0",
        f"Generated Timestamp:   {now_utc}",
        f"Restaurant Name:       {restaurant.name}",
        f"Legal Business Name:   {restaurant.legal_business_name if is_gst else 'N/A'}",
        f"GSTIN:                 {restaurant.gstin if is_gst else 'N/A (GST Disabled)'}",
        f"State / Code:          {restaurant.gst_state_name or 'N/A'} ({restaurant.gst_state_code or 'N/A'})",
        f"Restaurant Timezone:   Asia/Kolkata",
        f"GST Mode:              {'GST Enabled' if is_gst else 'GST Disabled'}",
        f"Reporting Period:      {preset or 'custom'} ({start_local.isoformat()} to {end_local.isoformat()})",
        "--------------------------------------------------------------------------------",
        "MANDATORY DISCLAIMER:",
        f"  {DISCLAIMER_TEXT}",
        "================================================================================",
    ]
    return "\n".join(lines)


def _apply_workbook_styles(ws: openpyxl.worksheet.worksheet.Worksheet):
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.border = thin_border
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)

        ws.column_dimensions[col_letter].width = max(14, min(max_len + 4, 40))


def build_sales_register_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    headers = [
        "Invoice Date", "Doc #", "Invoice #", "Doc Type", "Tax Type",
        "Customer Name", "Customer GSTIN", "Subtotal", "Discount",
        "Taxable Amount", "GST Rate (%)", "CGST", "SGST", "IGST",
        "Total Tax", "Total Amount", "Payment Status"
    ]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            subtotal = _to_num(r.get("subtotal"))
            discount = _to_num(r.get("discount_amount"))
            taxable = _to_num(r.get("taxable_amount"))
            gst_rate = _to_num(r.get("gst_rate"))
            cgst = _to_num(r.get("cgst_amount"))
            sgst = _to_num(r.get("sgst_amount"))
            igst = _to_num(r.get("igst_amount"))
            tax = _to_num(r.get("tax_amount"))
            total = _to_num(r.get("total_amount"))

            row = [
                r.get("invoice_date", "")[:10] if r.get("invoice_date") else "",
                r.get("document_number", ""),
                r.get("invoice_number") or "N/A",
                _map_val(r.get("document_type")),
                _map_val(r.get("customer_tax_type")),
                r.get("customer_legal_name") or "—",
                r.get("customer_gstin") or "—",
                subtotal, discount, taxable, gst_rate,
                cgst, sgst, igst, tax, total,
                _map_val(r.get("payment_status")),
            ]
            ws.append(row)

            # Apply numeric number formatting to monetary columns
            row_idx = ws.max_row
            for col_idx in [8, 9, 10, 12, 13, 14, 15, 16]:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "#,##0.00"
            ws.cell(row=row_idx, column=11).number_format = "0.00"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_gst_summary_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Rate Summary"

    headers = [
        "GST Rate (%)", "Sale Type", "Taxable Sales",
        "CGST Amount", "SGST Amount", "IGST Amount", "Total GST", "Document Count"
    ]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            rate = _to_num(r.get("gst_rate"))
            taxable = _to_num(r.get("taxable_amount"))
            cgst = _to_num(r.get("cgst_amount"))
            sgst = _to_num(r.get("sgst_amount"))
            igst = _to_num(r.get("igst_amount"))
            total_gst = _to_num(r.get("total_gst"))
            doc_cnt = r.get("document_count", 0)

            row = [rate, _map_val(r.get("customer_tax_type")), taxable, cgst, sgst, igst, total_gst, doc_cnt]
            ws.append(row)

            row_idx = ws.max_row
            ws.cell(row=row_idx, column=1).number_format = "0.00"
            for col_idx in [3, 4, 5, 6, 7]:
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=8).number_format = "#,##0"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_hsn_summary_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HSN SAC Summary"

    ws.append(["NOTE: " + data.get("tax_allocation_notice", DISCLAIMER_TEXT)])
    ws.append([])

    headers = ["HSN / SAC Code", "Item Description", "Total Quantity", "Line Item Count", "GST Rates Observed", "Taxable Amount"]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            rates_str = ", ".join(f"{x}%" for x in r.get("gst_rates_used", []))
            row = [
                r.get("hsn_sac_code"),
                r.get("description"),
                r.get("total_quantity"),
                r.get("line_count"),
                rates_str or "—",
                "Unallocated (Header Discount)",
            ]
            ws.append(row)

            row_idx = ws.max_row
            ws.cell(row=row_idx, column=3).number_format = "#,##0"
            ws.cell(row=row_idx, column=4).number_format = "#,##0"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_b2b_register_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B2B Register"

    headers = [
        "Invoice Date", "Invoice #", "Customer GSTIN", "Customer Legal Name",
        "Subtotal", "Taxable Amount", "CGST", "SGST", "IGST", "Total Amount"
    ]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            subtotal = _to_num(r.get("subtotal"))
            taxable = _to_num(r.get("taxable_amount"))
            cgst = _to_num(r.get("cgst_amount"))
            sgst = _to_num(r.get("sgst_amount"))
            igst = _to_num(r.get("igst_amount"))
            total = _to_num(r.get("total_amount"))

            row = [
                r.get("invoice_date", "")[:10] if r.get("invoice_date") else "",
                r.get("invoice_number") or r.get("document_number"),
                r.get("customer_gstin") or "—",
                r.get("customer_legal_name") or "B2B Customer",
                subtotal, taxable, cgst, sgst, igst, total
            ]
            ws.append(row)

            row_idx = ws.max_row
            for col_idx in [5, 6, 7, 8, 9, 10]:
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_b2c_register_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B2C Register"

    headers = [
        "Invoice Date", "Doc #", "Invoice #", "Doc Type",
        "Subtotal", "Discount", "Taxable Amount", "Total Tax", "Total Amount"
    ]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            subtotal = _to_num(r.get("subtotal"))
            discount = _to_num(r.get("discount_amount"))
            taxable = _to_num(r.get("taxable_amount"))
            tax = _to_num(r.get("tax_amount"))
            total = _to_num(r.get("total_amount"))

            row = [
                r.get("invoice_date", "")[:10] if r.get("invoice_date") else "",
                r.get("document_number"),
                r.get("invoice_number") or "—",
                _map_val(r.get("document_type")),
                subtotal, discount, taxable, tax, total
            ]
            ws.append(row)

            row_idx = ws.max_row
            for col_idx in [5, 6, 7, 8, 9]:
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_documents_issued_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documents Issued Audit"

    audit = data.get("audit", {})
    gaps = audit.get("sequence_gaps", [])
    if gaps:
        ws.append([f"NOTE: Detected {len(gaps)} sequence gap(s) requiring operational review."])
        ws.append([])

    headers = ["Invoice Number", "Invoice Date", "Doc Type", "Doc Number", "Status", "Sequence Note"]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            status_str = "Cancelled" if r.get("is_cancelled") else _map_val(r.get("status"))
            row = [
                r.get("invoice_number"),
                r.get("invoice_date", "")[:10] if r.get("invoice_date") else "",
                _map_val(r.get("document_type")),
                r.get("document_number"),
                status_str,
                "Consumed Sequence Number" if r.get("is_cancelled") else "Active Sequence"
            ]
            ws.append(row)

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cancelled_documents_xlsx(data: Dict[str, Any], restaurant: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cancelled Documents"

    ws.append(["NOTE: " + data.get("note", DISCLAIMER_TEXT)])
    ws.append([])

    headers = [
        "Created Date", "Bill #", "Invoice #", "Subtotal", "Discount",
        "Taxable Amount", "CGST", "SGST", "IGST", "Total Amount", "Cancellation Status"
    ]
    ws.append(headers)

    records = data.get("records", [])
    if not records:
        ws.append(["No transactions for selected period"] + [""] * (len(headers) - 1))
    else:
        for r in records:
            subtotal = _to_num(r.get("subtotal"))
            discount = _to_num(r.get("discount_amount"))
            taxable = _to_num(r.get("taxable_amount"))
            cgst = _to_num(r.get("cgst_amount"))
            sgst = _to_num(r.get("sgst_amount"))
            igst = _to_num(r.get("igst_amount"))
            total = _to_num(r.get("total_amount"))

            row = [
                r.get("created_at", "")[:10] if r.get("created_at") else "",
                r.get("document_number"),
                r.get("invoice_number") or "—",
                subtotal, discount, taxable, cgst, sgst, igst, total,
                _map_val(r.get("cancellation_status"))
            ]
            ws.append(row)

            row_idx = ws.max_row
            for col_idx in [4, 5, 6, 7, 8, 9, 10]:
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    _apply_workbook_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_export(records: List[Dict[str, Any]], field_keys: List[Tuple[str, str]]) -> str:
    output = io.StringIO()
    # Headers
    output.write(",".join(f'"{h}"' for _, h in field_keys) + "\n")

    if not records:
        output.write('"No transactions for selected period"' + "," * (len(field_keys) - 1) + "\n")
    else:
        for r in records:
            row_vals = []
            for k, _ in field_keys:
                val = r.get(k)
                if val is None:
                    row_vals.append('""')
                else:
                    mapped = _map_val(val)
                    row_vals.append(f'"{mapped}"')
            output.write(",".join(row_vals) + "\n")

    return output.getvalue()


def build_gst_summary_pdf(data: Dict[str, Any], restaurant: Any) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=36, rightMargin=36,
        topMargin=36, bottomMargin=36
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DocTitle", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#1F2937")
    )
    sub_style = ParagraphStyle(
        "DocSub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9, leading=12, textColor=colors.HexColor("#4B5563")
    )
    section_style = ParagraphStyle(
        "SectionHead", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=12, leading=16, textColor=colors.HexColor("#111827")
    )

    is_gst = bool(data.get("gst_enabled"))
    summary = data.get("summary", {})
    period = data.get("period", {})

    elements = []

    # Header Title
    elements.append(Paragraph(f"OMLU GST Summary Report — {restaurant.name}", title_style))
    elements.append(Spacer(1, 4))

    meta_str = f"GSTIN: {data.get('gstin') or 'N/A (GST Disabled)'} | Period: {period.get('start_date')} to {period.get('end_date')} | Timezone: Asia/Kolkata"
    elements.append(Paragraph(meta_str, sub_style))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#D1D5DB"), spaceAfter=12))

    # KPI Summary Table
    kpi_data = [
        ["Gross Sales", "Discounts", "Taxable Sales", "Total GST", "Net Sales", "Doc Count"],
        [
            f"INR {summary.get('gross_sales', '0.00')}",
            f"INR {summary.get('discount_amount', '0.00')}",
            f"INR {summary.get('taxable_sales', '0.00')}" if is_gst else "N/A",
            f"INR {summary.get('total_gst', '0.00')}" if is_gst else "N/A",
            f"INR {summary.get('net_sales', '0.00')}",
            str(summary.get("document_count", 0)),
        ]
    ]
    t_kpi = Table(kpi_data, colWidths=[120, 120, 120, 120, 120, 80])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 14))

    # Tax Components
    if is_gst:
        elements.append(Paragraph("Tax Component Breakdown", section_style))
        elements.append(Spacer(1, 6))
        tax_data = [
            ["CGST (Central Tax)", "SGST (State Tax)", "IGST (Integrated Tax)", "Total GST"],
            [
                f"INR {summary.get('cgst_amount', '0.00')}",
                f"INR {summary.get('sgst_amount', '0.00')}",
                f"INR {summary.get('igst_amount', '0.00')}",
                f"INR {summary.get('total_gst', '0.00')}",
            ]
        ]
        t_tax = Table(tax_data, colWidths=[180, 180, 180, 180])
        t_tax.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ECFDF5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#065F46')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#A7F3D0')),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t_tax)
        elements.append(Spacer(1, 14))

    # Disclaimer Footer
    disc_style = ParagraphStyle(
        "Disclaimer", parent=styles["Normal"],
        fontName="Helvetica-Oblique", fontSize=8, leading=10, textColor=colors.HexColor("#6B7280")
    )
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"NOTE: {DISCLAIMER_TEXT}", disc_style))

    doc.build(elements)
    return buf.getvalue()


def generate_ca_package_zip(
    db: Any,
    staff: StaffUser,
    preset: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Tuple[bytes, str]:
    restaurant = staff.restaurant
    start_local, end_local, _, _ = resolve_gst_period_bounds(restaurant, preset, start_date, end_date)
    is_gst = bool(restaurant.gst_enabled)

    slug = sanitize_filename(restaurant.slug or restaurant.name)
    period_str = f"{start_local.isoformat()}_to_{end_local.isoformat()}"
    folder_name = f"OMLU_GST_{slug}_{period_str}"
    zip_filename = f"{folder_name}.zip"

    summary_data = get_gst_center_summary(db, staff, preset, start_date, end_date)
    sales_data = get_gst_sales_register(db, staff, preset, start_date, end_date, page=1, limit=10000)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1. report-metadata.txt
        meta_txt = generate_report_metadata(restaurant, preset, start_local, end_local)
        zf.writestr(f"{folder_name}/report-metadata.txt", meta_txt)

        # 2. sales-register.xlsx
        zf.writestr(f"{folder_name}/sales-register.xlsx", build_sales_register_xlsx(sales_data, restaurant))

        if is_gst:
            rate_data = get_gst_rate_summary(db, staff, preset, start_date, end_date)
            hsn_data = get_hsn_summary(db, staff, preset, start_date, end_date)
            b2b_data = get_b2b_register(db, staff, preset, start_date, end_date, page=1, limit=10000)
            b2c_data = get_b2c_register(db, staff, preset, start_date, end_date, page=1, limit=10000)
            audit_data = get_documents_issued_audit(db, staff, preset, start_date, end_date)
            cancelled_data = get_cancelled_documents_register(db, staff, preset, start_date, end_date, page=1, limit=10000)

            zf.writestr(f"{folder_name}/gst-summary.xlsx", build_gst_summary_xlsx(rate_data, restaurant))
            zf.writestr(f"{folder_name}/hsn-sac-summary.xlsx", build_hsn_summary_xlsx(hsn_data, restaurant))
            zf.writestr(f"{folder_name}/b2b-invoices.xlsx", build_b2b_register_xlsx(b2b_data, restaurant))
            zf.writestr(f"{folder_name}/b2c-summary.xlsx", build_b2c_register_xlsx(b2c_data, restaurant))
            zf.writestr(f"{folder_name}/documents-issued.xlsx", build_documents_issued_xlsx(audit_data, restaurant))
            zf.writestr(f"{folder_name}/cancelled-documents.xlsx", build_cancelled_documents_xlsx(cancelled_data, restaurant))
            zf.writestr(f"{folder_name}/gst-summary.pdf", build_gst_summary_pdf(summary_data, restaurant))
        else:
            # GST Disabled -> include summary PDF & sales register
            zf.writestr(f"{folder_name}/gst-summary.pdf", build_gst_summary_pdf(summary_data, restaurant))

    return buf.getvalue(), zip_filename
